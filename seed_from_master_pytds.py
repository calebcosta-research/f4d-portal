"""Load ALL trust funds from F4D_rd_master.xlsx into WBG.<schema> on SQL Server.

Pure-Python loader for the locked-down WB VDI. It reads the master workbook with
**openpyxl** and writes with **python-tds** (``import pytds``) -- the only DB tools
that import under the VDI's group-policy DLL block. No pandas, pyodbc, or
SQLAlchemy (all of those pull in blocked compiled extensions).

It reproduces seed_all_from_master.py exactly: for each Trust Fund Number it
creates a Team, a User (username=<TFnum>_admin, password=<TFnum>_p), a TrustFund
(name == username), the deliverable/result Indicators + Mandatory mappings, and a
filled FY25 report (basic info, objective, deliverable and results blobs stored
as ``str(dict)`` just like the app expects).

Install deps (no admin needed):
    py -m pip install python-tds openpyxl

Credentials come from the environment (never hard-coded here):
    $env:sql_username="EFI_Admin"; $env:sql_password="<password>"
...or the script prompts for them. Server/db/port/schema default to the known
values and can be overridden by env (sql_host, sql_database, sql_port, db_schema).

Run:
    $env:F4D_MASTER="C:\\Users\\wb620297\\f4d\\F4D_rd_master.xlsx"
    py seed_from_master_pytds.py

Dry run (parse + report only, no DB, no driver needed):
    $env:F4D_DRYRUN="1"; py seed_from_master_pytds.py

Re-runnable: each trust fund commits in its own transaction and is skipped if its
user already exists, so a re-run resumes cleanly after any partial failure.
"""
import datetime
import getpass
import os
from collections import defaultdict

from openpyxl import load_workbook

MASTER = os.environ.get("F4D_MASTER", r"C:\Users\wb620297\f4d\F4D_rd_master.xlsx")
SCHEMA = os.environ.get("db_schema", "f4d")
DRY = os.environ.get("F4D_DRYRUN") == "1"
# Full-refresh guard: ONLY when F4D_WIPE == "YES" does the loader DELETE every
# existing f4d.* row before reloading, making the database an exact mirror of the
# workbook. Any other value (or unset) leaves existing data and does a normal
# additive load (new funds added, existing ones skipped).
WIPE = os.environ.get("F4D_WIPE") == "YES"
# Two projects were each issued two trust-fund numbers (a clerical oddity). The
# alias number is loaded as a second LOGIN attached to the primary fund's team
# (no separate fund/report), so both logins open the same submission. Mirrors
# context.py _TF_ALIASES. Format: {alias_tfnum: primary_tfnum}.
ALIAS_PRIMARY = {"TF0C8998": "TF0C8995", "TF0C8493": "TF0C8491"}

NOW = datetime.datetime.now()
TARGET_FY = "FY25"
FISCAL_YEARS = ("FY24", "FY25", "FY26")
CATS = "Completed, In Progress, Not Started"
DELIV_CATS = ("Completed", "In Progress", "Not Started")
F4D_ASSOC = "Yes, this P code is used solely for F4D funded activities"

PILLAR_NAMES = {
    "pillar 1": "Pillar 1: Strengthening Financial Sector Resiliency",
    "pillar 2": "Pillar 2: Financing the Poor and Vulnerable",
    "pillar 3": "Pillar 3: Financing the Real Economy",
    "pillar 4": "Pillar 4: Developing Financial Markets",
}

# Table DDL, generated from model.py via SQLAlchemy's SQL Server compiler (so it
# is byte-for-byte what the app's create_all() would build) and listed in
# dependency order. Each is applied only if the table is missing, so this is
# idempotent and coexists with the app's own create_all(). {s} = schema.
_DDL = [
    ("countries", """CREATE TABLE {s}.countries (
        id INTEGER NOT NULL IDENTITY, country VARCHAR(max) NOT NULL,
        deleted BIT NULL, created_at DATETIME NOT NULL, updated_at DATETIME NOT NULL,
        PRIMARY KEY (id))"""),
    ("fys", """CREATE TABLE {s}.fys (
        id INTEGER NOT NULL IDENTITY, fy VARCHAR(max) NOT NULL,
        deleted BIT NULL, created_at DATETIME NOT NULL, updated_at DATETIME NOT NULL,
        PRIMARY KEY (id))"""),
    ("regions", """CREATE TABLE {s}.regions (
        id INTEGER NOT NULL IDENTITY, region VARCHAR(max) NOT NULL,
        deleted BIT NULL, created_at DATETIME NOT NULL, updated_at DATETIME NOT NULL,
        PRIMARY KEY (id))"""),
    ("teams", """CREATE TABLE {s}.teams (
        id INTEGER NOT NULL IDENTITY, team VARCHAR(max) NOT NULL,
        deleted BIT NULL, created_at DATETIME NOT NULL, updated_at DATETIME NOT NULL,
        PRIMARY KEY (id))"""),
    ("indicators", """CREATE TABLE {s}.indicators (
        id INTEGER NOT NULL IDENTITY, indicator_id VARCHAR(max) NOT NULL,
        parent_id VARCHAR(max) NULL, indicator_name TEXT NOT NULL,
        standard_indicator_name TEXT NULL, indicator_prompt TEXT NULL,
        pillar_info VARCHAR(max) NULL, tier_info VARCHAR(max) NULL,
        indicator_definition TEXT NULL, unit_of_measurement VARCHAR(11) NULL,
        categorical_unit TEXT NULL, indicator_conversion VARCHAR(max) NULL,
        custom_indicator BIT NULL, team_id INTEGER NULL, deleted BIT NULL,
        created_at DATETIME NOT NULL, updated_at DATETIME NOT NULL,
        PRIMARY KEY (id), FOREIGN KEY(team_id) REFERENCES {s}.teams (id))"""),
    ("trustfunds", """CREATE TABLE {s}.trustfunds (
        id INTEGER NOT NULL IDENTITY, project_type VARCHAR(10) NOT NULL,
        name VARCHAR(max) NOT NULL, description VARCHAR(max) NULL,
        pcode VARCHAR(max) NULL, [grant] VARCHAR(max) NULL, ttl VARCHAR(max) NULL,
        team_id INTEGER NULL, deleted BIT NULL, created_at DATETIME NOT NULL,
        updated_at DATETIME NOT NULL, PRIMARY KEY (id),
        FOREIGN KEY(team_id) REFERENCES {s}.teams (id))"""),
    ("users", """CREATE TABLE {s}.users (
        id INTEGER NOT NULL IDENTITY, username VARCHAR(255) NOT NULL,
        password VARCHAR(255) NOT NULL, team_id INTEGER NULL, deleted BIT NULL,
        created_at DATETIME NOT NULL, updated_at DATETIME NOT NULL,
        PRIMARY KEY (id), UNIQUE (username),
        FOREIGN KEY(team_id) REFERENCES {s}.teams (id))"""),
    ("grant_info_long", """CREATE TABLE {s}.grant_info_long (
        id INTEGER NOT NULL IDENTITY, trustfund_id INTEGER NOT NULL,
        fiscal_year_id INTEGER NULL, field VARCHAR(max) NOT NULL, value TEXT NULL,
        team_id INTEGER NULL, deleted BIT NULL, created_at DATETIME NOT NULL,
        updated_at DATETIME NOT NULL, PRIMARY KEY (id),
        FOREIGN KEY(trustfund_id) REFERENCES {s}.trustfunds (id),
        FOREIGN KEY(fiscal_year_id) REFERENCES {s}.fys (id),
        FOREIGN KEY(team_id) REFERENCES {s}.teams (id))"""),
    ("trustfund_indicator_mapping", """CREATE TABLE {s}.trustfund_indicator_mapping (
        id INTEGER NOT NULL IDENTITY, trustfund_id INTEGER NOT NULL,
        indicator_id INTEGER NOT NULL, relation_ship VARCHAR(9) NOT NULL,
        team_id INTEGER NULL, deleted BIT NULL, created_at DATETIME NOT NULL,
        updated_at DATETIME NOT NULL, PRIMARY KEY (id),
        FOREIGN KEY(trustfund_id) REFERENCES {s}.trustfunds (id),
        FOREIGN KEY(indicator_id) REFERENCES {s}.indicators (id),
        FOREIGN KEY(team_id) REFERENCES {s}.teams (id))"""),
]


# --------------------------------------------------------------------------- #
# Small helpers (ported verbatim from seed_all_from_master.py)                 #
# --------------------------------------------------------------------------- #
def norm(v):
    if v is None or (isinstance(v, float) and v != v):
        return ""
    return str(v).strip()


def to_float(v):
    try:
        return float(norm(v).replace(",", "").replace("$", "").replace("%", ""))
    except (ValueError, AttributeError):
        return None


def progress_category(v):
    v = norm(v)
    if v.lower() in ("complete", "completed"):
        return "Completed"
    if v in DELIV_CATS:
        return v
    return ""


def results_unit_value(unit, raw):
    u = norm(unit).lower()
    f = to_float(raw)
    if u in ("number", "us$", "count") and f is not None:
        return "Number", f
    if "percent" in u and f is not None:
        return "Percentage", f
    return "Short Text", norm(raw)


def find_col(cols, *subs):
    for c in cols:
        cl = str(c).lower()
        if all(s.lower() in cl for s in subs):
            return c
    return None


# --------------------------------------------------------------------------- #
# Excel reading (openpyxl, replaces pandas.read_excel)                         #
# --------------------------------------------------------------------------- #
def read_sheet(path, sheet):
    """Return (headers, rows) where each row is a dict header -> raw cell value."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    headers = [norm(h) for h in next(it)]
    rows = []
    for raw in it:
        if raw is None or all(c is None for c in raw):
            continue
        rows.append({headers[i]: (raw[i] if i < len(raw) else None)
                     for i in range(len(headers))})
    wb.close()
    return headers, rows


# --------------------------------------------------------------------------- #
# Parse: build one structured record per trust fund (no DB, fully testable)    #
# --------------------------------------------------------------------------- #
def parse(path):
    dcols, deliv = read_sheet(path, "Deliverables")
    rcols, res = read_sheet(path, "Results")

    D = {"tf": "Trust Fund Number", "name": "Deliverable name",
         "prog": "Progress_clean", "desc": find_col(dcols, "brief description"),
         "src": "Data source", "media": find_col(dcols, "photos"),
         "pub": find_col(dcols, "publicly available"), "std": "standard_indicator"}
    R = {"tf": "Trust Fund Number", "name": "Indicator name",
         "unit": "Unit of measure", "prog": "Progress_clean",
         "base": find_col(rcols, "baseline value"),
         "target": find_col(rcols, "target value"),
         "how": find_col(rcols, "how are you collecting"), "std": "standard_indicator"}
    META = {"pcode": find_col(dcols, "project pcode") or find_col(dcols, "pcode"),
            "region": "Region", "country": "Country",
            "title": find_col(dcols, "project title"),
            "ttl": find_col(dcols, "f4d ttl name"), "objective": "Objective",
            "grant": "Grant Name"}
    PILLAR_COLS = {k: find_col(dcols, k) for k in PILLAR_NAMES}
    THEME_COLS = {
        "Climate change and sustainable finance": find_col(dcols, "climate change"),
        "Advancing digitalization": find_col(dcols, "digital financial"),
        "Financing solutions to close gender gaps": find_col(dcols, "gender")}

    deliv_by = defaultdict(list)
    res_by = defaultdict(list)
    for row in deliv:
        deliv_by[norm(row.get(D["tf"]))].append(row)
    for row in res:
        res_by[norm(row.get(R["tf"]))].append(row)
    all_tfs = sorted(t for t in set(deliv_by) | set(res_by) if t)

    records = []
    for tfnum in all_tfs:
        drows = deliv_by.get(tfnum, [])
        rrows = res_by.get(tfnum, [])
        meta = drows[0] if drows else rrows[0]

        delivs = []
        i = 0
        for row in drows:
            dname = norm(row.get(D["name"]))
            if not dname:
                continue
            i += 1
            has_media = "Yes" if (norm(row.get(D["media"])) or norm(row.get(D["pub"]))) else "No"
            delivs.append({
                "indicator_id": f"{tfnum}-D{i}", "name": dname,
                "prompt": "Progress Update", "unit": "Categorical",
                "categorical": CATS, "std": norm(row.get(D["std"])),
                "blob": {"input_value": progress_category(row.get(D["prog"])),
                         "progress": "", "deliverable_quantity": "",
                         "description": norm(row.get(D["desc"])),
                         "data_source": norm(row.get(D["src"])),
                         "next_steps": "", "supporting_materials_url": has_media}})

        results = []
        j = 0
        for row in rrows:
            rname = norm(row.get(R["name"]))
            if not rname:
                continue
            j += 1
            unit, ival = results_unit_value(row.get(R["unit"]), row.get(R["prog"]))
            results.append({
                "indicator_id": f"{tfnum}-RI{j}", "name": rname,
                "prompt": rname, "unit": unit, "std": norm(row.get(R["std"])),
                "blob": {"input_value": ival, "baseline_value": norm(row.get(R["base"])),
                         "year_baseline": None, "progress": "",
                         "target_value": norm(row.get(R["target"])), "year_target": None,
                         "data_collection": norm(row.get(R["how"])),
                         "level_of_result": "Output"}})

        pillars = [PILLAR_NAMES[k] for k, c in PILLAR_COLS.items()
                   if c and norm(meta.get(c)) in ("1", "1.0", "yes", "y")]
        ccts = [name for name, c in THEME_COLS.items()
                if c and norm(meta.get(c)) in ("1", "1.0", "yes", "y")]

        records.append({
            "tfnum": tfnum,
            "pcode": norm(meta.get(META["pcode"])),
            "description": norm(meta.get(META["title"])) or norm(meta.get(META["grant"])),
            "ttl": norm(meta.get(META["ttl"])),
            "region": norm(meta.get(META["region"])).title(),
            "country": norm(meta.get(META["country"])),
            "pillars": ", ".join(pillars), "ccts": ", ".join(ccts),
            "objective": norm(meta.get(META["objective"])),
            "delivs": delivs, "results": results})
    return records


# --------------------------------------------------------------------------- #
# DB write helpers (raw pytds, parameterised)                                  #
# --------------------------------------------------------------------------- #
def insert_id(cur, table, cols, vals):
    ph = ", ".join(["%s"] * len(vals))
    cur.execute(f"INSERT INTO {SCHEMA}.{table} ({', '.join(cols)}) "
                f"OUTPUT INSERTED.id VALUES ({ph})", tuple(vals))
    return cur.fetchone()[0]


def insert_row(cur, table, cols, vals):
    ph = ", ".join(["%s"] * len(vals))
    cur.execute(f"INSERT INTO {SCHEMA}.{table} ({', '.join(cols)}) VALUES ({ph})",
                tuple(vals))


def region_id_for(cur, name):
    if not name:
        return None
    cur.execute(f"SELECT id FROM {SCHEMA}.regions WHERE region=%s", (name,))
    r = cur.fetchone()
    if r:
        return r[0]
    return insert_id(cur, "regions", ["region", "deleted", "created_at", "updated_at"],
                     [name, 0, NOW, NOW])


def ensure_country(cur, name):
    if not name:
        return
    cur.execute(f"SELECT id FROM {SCHEMA}.countries WHERE country=%s", (name,))
    if not cur.fetchone():
        insert_row(cur, "countries", ["country", "deleted", "created_at", "updated_at"],
                   [name, 0, NOW, NOW])


def ensure_fy(cur, label):
    cur.execute(f"SELECT id FROM {SCHEMA}.fys WHERE fy=%s", (label,))
    r = cur.fetchone()
    if r:
        return r[0]
    return insert_id(cur, "fys", ["fy", "deleted", "created_at", "updated_at"],
                     [label, 0, NOW, NOW])


def load_record(cur, rec, fy25_id):
    tfnum = rec["tfnum"]
    username = f"{tfnum}_admin"

    team_id = insert_id(cur, "teams", ["team", "deleted", "created_at", "updated_at"],
                        [f"Team {tfnum}", 0, NOW, NOW])
    insert_row(cur, "users",
               ["username", "password", "team_id", "deleted", "created_at", "updated_at"],
               [username, f"{tfnum}_p", team_id, 0, NOW, NOW])
    tf_id = insert_id(cur, "trustfunds",
                      ["project_type", "name", "description", "pcode", "[grant]", "ttl",
                       "team_id", "deleted", "created_at", "updated_at"],
                      ["Trust Fund", username, rec["description"], rec["pcode"], tfnum,
                       rec["ttl"], team_id, 0, NOW, NOW])

    ensure_country(cur, rec["country"])
    rid = region_id_for(cur, rec["region"])

    def add_ind(item, custom):
        ind_id = insert_id(cur, "indicators",
                           ["indicator_id", "indicator_name", "indicator_prompt",
                            "unit_of_measurement", "categorical_unit",
                            "standard_indicator_name", "custom_indicator", "team_id",
                            "deleted", "created_at", "updated_at"],
                           [item["indicator_id"], item["name"], item["prompt"],
                            item["unit"], item.get("categorical"),
                            item["std"] or None, 1 if custom else 0, team_id, 0, NOW, NOW])
        insert_row(cur, "trustfund_indicator_mapping",
                   ["trustfund_id", "indicator_id", "relation_ship", "team_id",
                    "deleted", "created_at", "updated_at"],
                   [tf_id, ind_id, "Mandatory", team_id, 0, NOW, NOW])
        return ind_id

    deliv_blob = {}
    for item in rec["delivs"]:
        deliv_blob[str(add_ind(item, custom=False))] = item["blob"]
    res_blob = {}
    for item in rec["results"]:
        res_blob[str(add_ind(item, custom=True))] = item["blob"]

    def addf(field, value):
        if value:
            insert_row(cur, "grant_info_long",
                       ["trustfund_id", "fiscal_year_id", "field", "[value]", "team_id",
                        "deleted", "created_at", "updated_at"],
                       [tf_id, fy25_id, field, value, team_id, 0, NOW, NOW])

    addf("country", rec["country"])
    addf("region_id", str(rid) if rid is not None else None)
    addf("f4d_association", F4D_ASSOC)
    addf("pillars", rec["pillars"])
    addf("ccts", rec["ccts"])
    addf("strategic_objective", rec["objective"])
    if deliv_blob:
        addf("deliverables", str(deliv_blob))
    if res_blob:
        addf("custom_indicators", str(res_blob))


def create_tables(cur):
    """Create the f4d.* tables if they don't already exist (idempotent)."""
    for name, ddl in _DDL:
        cur.execute(f"IF OBJECT_ID('{SCHEMA}.{name}','U') IS NULL "
                    + ddl.format(s=SCHEMA))


# --------------------------------------------------------------------------- #
def wipe_f4d(cur):
    """Delete ALL rows from the f4d.* tables (children before parents) for a full
    refresh. Guarded by F4D_WIPE=YES. The schema and tables stay in place; only
    their data is removed, so the subsequent load rebuilds everything from scratch."""
    order = ["grant_info_long", "trustfund_indicator_mapping", "indicators",
             "trustfunds", "users", "fys", "regions", "countries", "teams"]
    for t in order:
        cur.execute(f"DELETE FROM {SCHEMA}.{t}")


def main():
    records = parse(MASTER)
    n_del = sum(len(r["delivs"]) for r in records)
    n_res = sum(len(r["results"]) for r in records)
    print(f"Parsed {len(records)} trust funds, {n_del} deliverables, {n_res} results "
          f"from {MASTER}")
    sample = records[0] if records else None
    if sample:
        print(f"  e.g. {sample['tfnum']}: pcode={sample['pcode']!r} "
              f"region={sample['region']!r} country={sample['country']!r} "
              f"delivs={len(sample['delivs'])} results={len(sample['results'])}")
    if WIPE:
        print("** WIPE MODE ON: the real run will DELETE all existing f4d.* rows "
              "first, then rebuild from the workbook above. **")
    if DRY:
        print("DRY RUN: parsed only, no database writes.")
        return

    import pytds
    user = os.environ.get("sql_username") or input("SQL username: ").strip()
    password = os.environ.get("sql_password") or getpass.getpass("SQL password: ")
    conn = pytds.connect(
        os.environ.get("sql_host", "WBGMSSQLEFIP001"),
        os.environ.get("sql_database", "WBG"), user, password,
        port=int(os.environ.get("sql_port", "5800")),
        autocommit=False, login_timeout=15)
    cur = conn.cursor()

    create_tables(cur)
    conn.commit()
    print(f"Tables ready in {SCHEMA}.*")

    if WIPE:
        wipe_f4d(cur)
        conn.commit()
        print("WIPED all existing rows from f4d.* — rebuilding from the workbook.")

    for label in FISCAL_YEARS:
        ensure_fy(cur, label)
    conn.commit()
    cur.execute(f"SELECT id FROM {SCHEMA}.fys WHERE fy=%s", (TARGET_FY,))
    fy25_id = cur.fetchone()[0]

    loaded = skipped = aliased = 0
    errors = []
    for rec in records:
        tfnum = rec["tfnum"]
        username = f"{tfnum}_admin"
        try:
            cur.execute(f"SELECT 1 FROM {SCHEMA}.users WHERE username=%s", (username,))
            if cur.fetchone():
                skipped += 1
                continue
            if tfnum in ALIAS_PRIMARY:
                # Alias number: add only a second login on the primary fund's team,
                # pointing at the same submission (no separate fund/report).
                primary_user = f"{ALIAS_PRIMARY[tfnum]}_admin"
                cur.execute(f"SELECT team_id FROM {SCHEMA}.trustfunds WHERE name=%s",
                            (primary_user,))
                row = cur.fetchone()
                if not row:
                    errors.append((tfnum, f"primary {ALIAS_PRIMARY[tfnum]} not loaded"))
                    continue
                insert_row(cur, "users",
                           ["username", "password", "team_id", "deleted", "created_at", "updated_at"],
                           [username, f"{tfnum}_p", row[0], 0, NOW, NOW])
                conn.commit()
                aliased += 1
                continue
            load_record(cur, rec, fy25_id)
            conn.commit()
            loaded += 1
        except Exception as exc:  # noqa: BLE001
            conn.rollback()
            errors.append((rec["tfnum"], str(exc)[:200]))
    conn.close()

    print(f"Loaded {loaded} funds, {aliased} alias logins (shared submission), "
          f"skipped {skipped}, {len(errors)} errors, of {len(records)} parsed.")
    print("Login scheme: username=<TFnum>_admin  password=<TFnum>_p")
    for t, e in errors[:15]:
        print(f"  ERROR {t}: {e}")


if __name__ == "__main__":
    main()
