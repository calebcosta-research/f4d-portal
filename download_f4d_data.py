"""Download all F4D Results Reporting data into a readable Excel workbook.

Read-only. Connects to the WB SQL Server, reads every grant's submissions, and
writes a single Excel file with clean, decoded sheets (Grant Info, Deliverables,
Results Indicators, Lending Operations, Collaboration, and a complete Raw
sheet). Nothing is written to the database.

Each grant's lending operations and collaborations also travel as columns on
the Deliverables and Results Indicators sheets, so a single row carries the
whole picture for pivot tables.

The M&E portfolio spreadsheet (window, donor, dates, grant status, pillars,
objective, ...) is merged onto the Deliverables and Results Indicators sheets,
matched to each portal entry by trust fund number. Point PORTFOLIO_DATA_PATH
(the first setting below) at your copy of that file.

Uses python-tds (``import pytds``) + openpyxl — the pure-Python tools that work on
the locked-down VDI and on a normal machine after a one-time
``pip install python-tds openpyxl``.

HOW TO RUN
  1. Make sure you can reach the WB database (on the WB network / VDI).
  2. Open a terminal in the folder that has this file.
  3. Run:   py download_f4d_data.py
     (on some machines the command is `python` instead of `py`)
  4. Answer the two questions it asks:
       * whose computer this is -- type your name if you are listed in
         KNOWN_USERS below, otherwise just press Enter;
       * the database password (Caleb will give it to you), then Enter.
  5. When it finishes it prints the name of the Excel file it created, in this
     same folder. Open that file in Excel.

If anything goes wrong the window stays open and explains what happened, and
writes the detail to download_error.txt next to this script -- send that file
on and it will say exactly what failed.

The portfolio spreadsheet is found in one of three ways, in this order:
PORTFOLIO_DATA_PATH below if you filled it in, then the saved location for
whoever answers the "whose computer is this" question, then any portfolio file
sitting next to this script. If none of them turn up a file the export still
works -- it just leaves the portfolio columns out and says so.

Re-downloading this script replaces the file, and with it anything you typed
into PORTFOLIO_DATA_PATH. Two ways to avoid re-typing it:
  * Keep the portfolio file in the same folder as this script (or the folder
    above it). It is found automatically and nothing needs editing.
  * Or set it once per terminal instead of editing the file, in PowerShell:
        $env:portfolio_data_path = "C:/Users/you/Documents/portfolio.csv"

Server / database / port default to the known F4D values. The username defaults
to EFI_Admin. All can be overridden with environment variables (sql_host,
sql_database, sql_port, sql_username, sql_password) but you normally don't need to.
"""
# =========================================================================== #
#  PORTFOLIO DATA FILE  --  EDIT THE LINE BELOW                               #
# =========================================================================== #
#  Put the location of the M&E portfolio file between the quotes, e.g.
#      PORTFOLIO_DATA_PATH = r"C:\Users\wb123456\Documents\F4D M&E Portfolio Data_MASTER(Project details).csv"
#
#  Tips
#    * Keep the  r  in front of the quotes on a Windows path.
#    * A folder also works -- the newest portfolio file in it is used.
#    * .csv and .xlsx both work.
#    * Leave it as ""  to look in this script's folder, the folder above it,
#      and the folder you are running from, for a file whose name contains
#      "portfolio" and "data".
#    * A one-off run can override it without editing:
#      set the portfolio_data_path environment variable instead.
#  Nothing is written to this file; it is only read.
PORTFOLIO_DATA_PATH = r""

#  Saved locations for people who can't edit this file. When PORTFOLIO_DATA_PATH
#  above is empty, the script asks whose computer it is running on and uses the
#  matching path from here. To add someone, copy a line and change the name and
#  the path (keep the r before the quotes).
KNOWN_USERS = {
    "Sara": r"C:\Users\wb293537\OneDrive - WBG\Desktop\F4D M&E Portfolio Data_MASTER.csv",
}

#  Which portfolio columns to add to the export. Empty list = all of them.
#  To keep the sheets narrow, list the exact headings you want instead, e.g.
#      PORTFOLIO_COLUMNS = ["Window Name", "Donor", "Grant status", "Closing Date"]
PORTFOLIO_COLUMNS = []

#  Headings in the portfolio file used to match its rows to the portal's grants.
#  The trust fund number is tried first, the P-code second.
PORTFOLIO_TF_COLUMN = "Trust Fund #"
PORTFOLIO_PCODE_COLUMN = "Project Pcode"
# =========================================================================== #

import ast
import csv
import datetime
import getpass
import glob
import os
import traceback

# Imported leniently so a missing library is reported by run() below, with the
# window held open, instead of vanishing when the console closes.
try:
    import pytds
except ImportError:
    pytds = None
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    Workbook = None

SCHEMA = os.environ.get("db_schema", "f4d")


def connect():
    user = os.environ.get("sql_username", "EFI_Admin")
    password = os.environ.get("sql_password") or getpass.getpass(
        "Enter the database password (ask Caleb), then press Enter: ")
    print("Connecting to the database ...")
    return pytds.connect(
        os.environ.get("sql_host", "WBGMSSQLEFIP001"),
        os.environ.get("sql_database", "WBG"), user, password,
        port=int(os.environ.get("sql_port", "5800")),
        autocommit=False, login_timeout=30)


def decode(value):
    """Turn a stored blob (str(dict)) into a real dict; {} if it isn't one."""
    if isinstance(value, str) and value.lstrip()[:1] in "{[":
        try:
            return ast.literal_eval(value)
        except (ValueError, SyntaxError):
            return {}
    return {}


# --------------------------------------------------------------------------- #
# Lending operations, CPFs and collaborations                                  #
# --------------------------------------------------------------------------- #
# The portal stores one numbered field per entry — operation_1, operation_2,
# cpf_1, collaboration_1, ... — each holding a str(dict). "collaborations"
# (no underscore) is the grant's single Yes/No answer, not an entry.
ENTRY_PREFIXES = ("operation_", "cpf_", "collaboration_")

# Columns the flattened lending-operations and collaboration info adds to the
# Deliverables and Results Indicators sheets, one grant-year per row.
SUMMARY_HEADERS = [
    "Lending Operations (#)", "Operation P Numbers", "Operation Names",
    "Operation Instruments", "Operation Approval FYs",
    "Total Commitment (US$m)", "Informed by F4D (US$m)", "Operation Evidence",
    "CPFs (#)", "CPF Countries", "CPF Years", "CPF Evidence",
    "Collaboration Reported?", "Collaborations (#)", "Collaboration Types",
    "Collaboration Partners", "Collaboration Description",
    "Collaboration Lessons Learned",
]


def entry_number(field):
    """Sort key for operation_2 / cpf_10 / collaboration_3."""
    tail = field.rsplit("_", 1)[-1]
    return int(tail) if tail.isdigit() else 0


def joined(entries, key):
    """Every non-empty value of one key across entries, in a single cell."""
    values = [as_text(e.get(key, "")).strip() for e in entries]
    return " | ".join(v for v in values if v)[:32000]


def number(value):
    """A money cell as a real number when it parses, so Excel can total it."""
    try:
        return float(str(value or "").replace(",", "").strip())
    except ValueError:
        return value


def total(entries, key):
    """Sum of a money/number key. Blank when nothing numeric was entered."""
    numbers = []
    for e in entries:
        try:
            numbers.append(float(str(e.get(key, "") or "").replace(",", "").strip()))
        except ValueError:
            continue
    return round(sum(numbers), 6) if numbers else ""


# --------------------------------------------------------------------------- #
# Portfolio data — read the spreadsheet and key it by trust fund number        #
# --------------------------------------------------------------------------- #
def match_key(value):
    """Normalise a trust fund number / P-code so 'tf0c1495 ' == 'TF0C1495'."""
    return "".join(ch for ch in str(value or "").upper() if ch.isalnum())


def column_letter(index):
    """0 -> A, 26 -> AA. Used to name portfolio columns that have no heading."""
    letters = ""
    index += 1
    while index:
        index, rem = divmod(index - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def ask_known_user():
    """Ask whose computer this is; return that person's saved path, or None.

    Skipped when nobody is listed in KNOWN_USERS, and when the script is run
    without a keyboard attached (a scheduled task), where input() gets no answer.
    """
    if not KNOWN_USERS:
        return None
    names = list(KNOWN_USERS)
    try:
        if len(names) == 1:
            answer = input(f"Is this {names[0]}? Type  yes  and press Enter. "
                           "If not, just press Enter: ")
            chosen = names[0] if answer.strip().lower() in ("y", "yes") else None
        else:
            answer = input(f"Whose computer is this? Type a name ({', '.join(names)}) "
                           "or just press Enter to skip: ")
            chosen = next((n for n in names
                           if n.lower() == answer.strip().lower()), None)
    except (EOFError, KeyboardInterrupt):
        return None
    if not chosen:
        return None
    path = KNOWN_USERS[chosen]
    if os.path.isfile(path):
        return path
    print(f"   ! {chosen}'s saved portfolio file isn't there any more:")
    print(f"     {path}")
    print("     Looking for it in the usual folders instead.")
    return None


def find_portfolio_file():
    """Locate the portfolio file. Returns a path, or None if there isn't one.

    Order: the portfolio_data_path environment variable, then
    PORTFOLIO_DATA_PATH at the top of this file, then an automatic search.
    """
    here = os.path.dirname(os.path.abspath(__file__))
    folders = [here, os.path.dirname(here), os.getcwd()]

    def newest_in(folder):
        hits = [f for ext in ("csv", "xlsx", "xlsm")
                for f in glob.glob(os.path.join(folder, "*." + ext))
                if "portfolio" in os.path.basename(f).lower()
                and "data" in os.path.basename(f).lower()
                and not os.path.basename(f).startswith("~$")]
        return max(hits, key=os.path.getmtime) if hits else None

    given = (os.environ.get("portfolio_data_path") or PORTFOLIO_DATA_PATH or "").strip().strip('"')
    if given:
        given = os.path.expanduser(given)
        if os.path.isdir(given):
            found = newest_in(given)
            if not found:
                print(f"   ! No portfolio file found in the folder: {given}")
            return found
        if os.path.isfile(given):
            return given
        # Just a file name was given: look for it in the usual folders.
        for folder in folders:
            if os.path.isfile(os.path.join(folder, given)):
                return os.path.join(folder, given)
        print(f"   ! PORTFOLIO_DATA_PATH points at something that isn't there: {given}")
        return None

    saved = ask_known_user()
    if saved:
        return saved

    for folder in folders:
        found = newest_in(folder)
        if found:
            return found
    return None


# Control characters Excel rejects, and its per-cell text limit.
ILLEGAL_IN_EXCEL = {c: None for c in list(range(9)) + [11, 12] + list(range(14, 32))}


def clean_cell(value):
    """Trimmed, Excel-safe text for one portfolio cell."""
    text = "" if value is None else str(value).strip()
    return text.translate(ILLEGAL_IN_EXCEL)[:32000]


def as_text(value):
    """Readable text for one stored value.

    Multi-select answers are kept as real Python lists inside the blob fields
    (a CPF's country is one), and openpyxl refuses a list outright. Render
    those as "Colombia, Peru" rather than "['Colombia', 'Peru']".
    """
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(v).strip() for v in value
                          if v is not None and str(v).strip())
    return "" if value is None else str(value)


def cell_for_excel(value):
    """One cell openpyxl will accept: numbers and dates as they are, everything
    else as text with lists flattened and control characters removed. A list,
    or a single stray character pasted into a text box, used to abort the
    whole export."""
    if value is None or isinstance(value, (int, float, bool,
                                           datetime.datetime, datetime.date)):
        return value
    return as_text(value).translate(ILLEGAL_IN_EXCEL)[:32000]


def add(worksheet, values):
    """Append one row, cell by cell, through cell_for_excel()."""
    worksheet.append([cell_for_excel(v) for v in values])


def read_table(path):
    """Read a .csv/.xlsx portfolio file into rows of text (first row = headings)."""
    if path.lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        ws = load_workbook(path, read_only=True, data_only=True).active
        return [[clean_cell(c) for c in row] for row in ws.iter_rows(values_only=True)]
    # Excel usually writes cp1252 here; utf-8 first in case it was saved that way.
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with open(path, newline="", encoding=encoding) as fh:
                return [[clean_cell(c) for c in row] for row in csv.reader(fh)]
        except UnicodeDecodeError:
            continue
    print(f"   ! Could not read {path} (unexpected file encoding).")
    return []


def load_portfolio():
    """(path, headings, {tf number: row}, {pcode: row}) for the portfolio file.

    Empty columns and the spreadsheet's trailing blank rows are dropped,
    repeated headings get a '(2)' suffix, and a cell listing several trust fund
    numbers ("TF0A1111, TF0A2222") is indexed under each of them.
    """
    path = find_portfolio_file()
    if not path:
        return None, [], {}, {}

    table = read_table(path)
    if len(table) < 2:
        print(f"   ! Portfolio file has no rows: {path}")
        return path, [], {}, {}

    raw_head, body = table[0], table[1:]
    body = [r for r in body if any(c for c in r)]

    # Keep columns that hold something. Unheaded columns keep their Excel letter.
    keep, headings, seen = [], [], {}
    for i, raw in enumerate(raw_head):
        if not any(i < len(r) and r[i] for r in body):
            continue  # empty column (the spreadsheet has thousands of them)
        name = raw or f"(column {column_letter(i)})"
        seen[name] = seen.get(name, 0) + 1
        if seen[name] > 1:
            name = f"{name} ({seen[name]})"
        keep.append(i)
        headings.append(name)

    wanted = [c for c in PORTFOLIO_COLUMNS if c]
    if wanted:
        missing = [c for c in wanted if c not in headings]
        if missing:
            print(f"   ! PORTFOLIO_COLUMNS not found in the file: {', '.join(missing)}")
        where = dict(zip(headings, keep))  # keep the order PORTFOLIO_COLUMNS lists
        keep = [where[c] for c in wanted if c in where]
        headings = [c for c in wanted if c in where]

    tf_col = next((i for i, h in zip(keep, headings) if h == PORTFOLIO_TF_COLUMN), None)
    pcode_col = next((i for i, h in zip(keep, headings) if h == PORTFOLIO_PCODE_COLUMN), None)
    # The join columns can be excluded from the output but still needed for matching.
    if tf_col is None:
        tf_col = next((i for i, h in enumerate(raw_head) if h == PORTFOLIO_TF_COLUMN), None)
    if pcode_col is None:
        pcode_col = next((i for i, h in enumerate(raw_head) if h == PORTFOLIO_PCODE_COLUMN), None)
    if tf_col is None and pcode_col is None:
        print(f"   ! Portfolio file has no '{PORTFOLIO_TF_COLUMN}' column — not merged.")
        return path, [], {}, {}

    by_tf, by_pcode = {}, {}
    for row in body:
        def cell(i):
            return row[i] if i is not None and i < len(row) else ""
        values = dict(zip(headings, [cell(i) for i in keep]))
        for number in str(cell(tf_col)).replace(";", ",").replace("/", ",").split(","):
            key = match_key(number)
            if key:
                by_tf.setdefault(key, values)
        key = match_key(cell(pcode_col))
        if key:
            by_pcode.setdefault(key, values)
    return path, headings, by_tf, by_pcode


def portfolio_keys(meta):
    """Trust fund numbers to try for one portal grant, best first.

    The portal stores the number in `grant`; `name` is the login
    ("TF0C1495_admin"), which still yields the number once the suffix is cut.
    """
    keys = []
    for value in (meta.get("grant"), meta.get("name")):
        key = match_key(value)
        if key.endswith("ADMIN"):
            key = key[:-5]
        if key and key not in keys:
            keys.append(key)
    return keys


def main():
    if pytds is None or Workbook is None:
        raise RuntimeError(
            "A library this script needs isn't installed. Run this one line, "
            "then try again:" + os.linesep + "    pip install python-tds openpyxl")

    pf_path, pf_headings, pf_by_tf, pf_by_pcode = load_portfolio()
    if pf_path and pf_headings:
        print(f"Portfolio data: {pf_path}")
        print(f"   {len(pf_headings)} columns, {len(pf_by_tf)} trust fund numbers.")
    elif not pf_path:
        print("Portfolio data: not found — the export will run without those columns.")
        print("   Set PORTFOLIO_DATA_PATH at the top of this file to add them.")

    conn = connect()
    cur = conn.cursor()

    # --- Look-up tables (id -> friendly text) ---------------------------------
    cur.execute(f"SELECT id, name, [grant], pcode, ttl, description FROM {SCHEMA}.trustfunds WHERE deleted=0")
    tf = {r[0]: {"name": r[1], "grant": r[2], "pcode": r[3], "ttl": r[4], "desc": r[5]}
          for r in cur.fetchall()}
    cur.execute(f"SELECT id, fy FROM {SCHEMA}.fys WHERE deleted=0")
    fy = {r[0]: r[1] for r in cur.fetchall()}
    cur.execute(f"SELECT id, indicator_name, unit_of_measurement FROM {SCHEMA}.indicators")
    ind = {r[0]: {"name": r[1], "unit": r[2]} for r in cur.fetchall()}
    cur.execute(f"SELECT id, region FROM {SCHEMA}.regions")
    region = {str(r[0]): r[1] for r in cur.fetchall()}

    def gname(tid):
        g = tf.get(tid, {})
        return g.get("grant") or g.get("name") or f"TrustFund {tid}"

    def iname(key, blob_entry):
        # admin-mapped indicators are keyed by numeric id; TTL-added ones carry a name.
        if isinstance(blob_entry, dict) and blob_entry.get("name"):
            return blob_entry["name"]
        try:
            return ind.get(int(key), {}).get("name") or f"Indicator {key}"
        except (TypeError, ValueError):
            return f"Indicator {key}"

    # Portfolio columns for a grant: matched on trust fund number, then P-code.
    # Headings that clash with a column the export already has are suffixed.
    taken = {"Trust Fund #", "Grant Name", "TTL", "Fiscal Year", "Country",
             "Region", "Product Line", "F4D Association", "Pillars",
             "Cross-Cutting Themes", "Strategic Objective", "Deliverable",
             "Indicator", "Unit", "Description", "Next Steps"}
    taken.update(SUMMARY_HEADERS)
    pf_out_headings = [f"{h} (portfolio)" if h in taken else h for h in pf_headings]
    pf_blank = [""] * len(pf_headings)
    pf_cache = {}
    pf_matched, pf_unmatched = set(), set()

    def portfolio(tid):
        """The portfolio row for one trust fund, as values in pf_headings order."""
        if tid not in pf_cache:
            meta = tf.get(tid, {})
            row = None
            for key in portfolio_keys(meta):
                row = pf_by_tf.get(key)
                if row:
                    break
            if row is None:
                row = pf_by_pcode.get(match_key(meta.get("pcode")))
            (pf_matched if row else pf_unmatched).add(gname(tid))
            pf_cache[tid] = [row.get(h, "") for h in pf_headings] if row else pf_blank
        return pf_cache[tid]

    # --- Read all submissions -------------------------------------------------
    print("Reading submissions ...")
    cur.execute(f"SELECT trustfund_id, fiscal_year_id, field, value, updated_at "
                f"FROM {SCHEMA}.grant_info_long WHERE deleted=0")
    rows = cur.fetchall()

    # Per-(grant, fiscal year) context so each Deliverables / Results row can
    # carry the grant's inputs (country, region, pillars, etc.) as own columns.
    CONTEXT = [
        ("country", "Country"), ("region_id", "Region"),
        ("p_code_instrument", "Product Line"), ("f4d_association", "F4D Association"),
        ("pillars", "Pillars"), ("ccts", "Cross-Cutting Themes"),
        ("strategic_objective", "Strategic Objective"),
    ]
    ctx_headers = [name for _, name in CONTEXT]
    ctx_fields = {f for f, _ in CONTEXT}
    context = {}
    for _tid, _fid, _field, _value, _upd in rows:
        if _field in ctx_fields:
            _v = region.get(str(_value), _value) if _field == "region_id" else _value
            context.setdefault((_tid, _fid), {})[_field] = "" if _v is None else _v

    def prefix(tid, fid):
        c = context.get((tid, fid), {})
        return ([tf.get(tid, {}).get("name", tid), gname(tid),
                 tf.get(tid, {}).get("ttl", ""), fy.get(fid, fid)]
                + [c.get(f, "") for f, _ in CONTEXT])

    # Lending operations, CPFs and collaborations, grouped per (grant, year)
    # and kept in the order the TTL entered them (operation_1, operation_2, ...).
    ops, cpfs, collabs, collab_answer = {}, {}, {}, {}
    for _tid, _fid, _field, _value, _upd in rows:
        if _field == "collaborations":
            collab_answer[(_tid, _fid)] = _value or ""
            continue
        if _field.startswith("operation_"):
            bucket = ops
        elif _field.startswith("cpf_"):
            bucket = cpfs
        elif _field.startswith("collaboration_"):
            bucket = collabs
        else:
            continue
        entry = decode(_value)
        if entry:
            bucket.setdefault((_tid, _fid), []).append((entry_number(_field), entry))
    for bucket in (ops, cpfs, collabs):
        for key in list(bucket):
            bucket[key] = [e for _, e in sorted(bucket[key], key=lambda pair: pair[0])]

    summary_cache = {}

    def summary(tid, fid):
        """The grant-year's operations and collaborations, as SUMMARY_HEADERS."""
        if (tid, fid) not in summary_cache:
            o = ops.get((tid, fid), [])
            c = cpfs.get((tid, fid), [])
            b = collabs.get((tid, fid), [])
            summary_cache[(tid, fid)] = [
                len(o) or "", joined(o, "p_number"), joined(o, "operation_name"),
                joined(o, "p_code_instrument"), joined(o, "approval_fy"),
                total(o, "total_commitment"), total(o, "informed_by_f4d"),
                joined(o, "evidence"),
                len(c) or "", joined(c, "country"), joined(c, "year"),
                joined(c, "evidence"),
                collab_answer.get((tid, fid), ""), len(b) or "",
                joined(b, "type"), joined(b, "partner_detail"),
                joined(b, "describe"), joined(b, "lessons_learned"),
            ]
        return summary_cache[(tid, fid)]

    # Per-grant progress for the current reporting year (the newest fiscal year):
    # Not Started = no data for it, In Progress = saved but not submitted,
    # Complete = a submission was recorded (report_submitted_at / report_status).
    newest_fy = max(fy.keys()) if fy else None
    prog = {}
    for _tid, _fid, _field, _value, _upd in rows:
        if _fid != newest_fy:
            continue
        p = prog.setdefault(_tid, {"last": None, "submitted_at": "", "complete": False})
        if _upd and (p["last"] is None or _upd > p["last"]):
            p["last"] = _upd
        if _field == "report_submitted_at" and _value:
            p["submitted_at"], p["complete"] = _value, True
        elif _field == "report_status" and _value:
            p["complete"] = True

    wb = Workbook()
    HEAD = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF")

    def new_sheet(title, headers):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            ws.cell(row=1, column=c).fill = HEAD
            ws.cell(row=1, column=c).font = hfont
            ws.cell(row=1, column=c).alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"
        return ws

    wb.remove(wb.active)  # drop default sheet

    # Project Progress — one row per grant, at-a-glance status for the reporting year.
    ws_prog = new_sheet("Project Progress",
        ["Trust Fund #", "Grant Name", "TTL", "Status", "Last Updated", "Submitted At"])
    for _tid, _meta in sorted(tf.items(), key=lambda kv: (kv[1].get("name") or "")):
        p = prog.get(_tid)
        if not p:
            add(ws_prog, [_meta.get("name"), gname(_tid), _meta.get("ttl", ""),
                            "Not Started", "", ""])
        else:
            add(ws_prog, [_meta.get("name"), gname(_tid), _meta.get("ttl", ""),
                            "Complete" if p["complete"] else "In Progress",
                            str(p["last"])[:19] if p["last"] else "", p["submitted_at"]])

    ws_info = new_sheet("Grant Info", [
        "Trust Fund #", "Grant Name", "TTL", "Fiscal Year", "Field", "Value"])
    ws_del = new_sheet("Deliverables",
        ["Trust Fund #", "Grant Name", "TTL", "Fiscal Year"] + ctx_headers +
        ["Deliverable", "Progress / Status", "Target #", "Number Completed",
         "Description", "Next Steps", "Photos/Materials?"] +
        SUMMARY_HEADERS + pf_out_headings)
    ws_res = new_sheet("Results Indicators",
        ["Trust Fund #", "Grant Name", "TTL", "Fiscal Year"] + ctx_headers +
        ["Indicator", "Unit", "Progress Value", "Explanation", "Baseline",
         "Baseline Yr", "Target", "Target Yr", "Level of Result",
         "Data Collection"] + SUMMARY_HEADERS + pf_out_headings)

    # One row per lending operation, then per CPF, for each grant and year.
    ws_ops = new_sheet("Lending Operations",
        ["Trust Fund #", "Grant Name", "TTL", "Fiscal Year"] + ctx_headers +
        ["Entry", "Entry #", "Informed?", "P Number", "Instrument",
         "Instrument (Other)", "Approval FY", "Operation Name",
         "Total Commitment (US$m)", "Informed by F4D (US$m)",
         "CPF Country", "CPF Year", "Evidence"])
    # One row per collaboration; grants that answered the question but listed
    # none still get a row, so the sheet accounts for every reporting grant.
    ws_collab = new_sheet("Collaboration",
        ["Trust Fund #", "Grant Name", "TTL", "Fiscal Year"] + ctx_headers +
        ["Collaboration Reported?", "Entry #", "Type",
         "Team(s) / Organization(s)", "Description", "Lessons Learned"])
    ws_raw = new_sheet("All Data (raw)", [
        "Trust Fund #", "Grant Name", "Fiscal Year", "Field", "Value", "Last Updated"])

    BLOB_FIELDS = {"deliverables", "custom_indicators"}

    for tid, fid, field, value, updated in rows:
        tfnum = tf.get(tid, {}).get("name", tid)
        grant = gname(tid)
        year = fy.get(fid, fid)

        # Raw catch-all (everything, unmodified)
        add(ws_raw, [tfnum, grant, year, field, value,
                     str(updated)[:19] if updated else ""])

        if field == "deliverables":
            for key, d in decode(value).items():
                if not isinstance(d, dict) or d.get("archived"):
                    continue
                add(ws_del, prefix(tid, fid) + [
                    iname(key, d), d.get("input_value", ""), d.get("progress", ""),
                    d.get("deliverable_quantity", ""), d.get("description", ""),
                    d.get("next_steps", ""), d.get("supporting_materials_url", "")]
                    + summary(tid, fid) + portfolio(tid))
        elif field == "custom_indicators":
            for key, d in decode(value).items():
                if not isinstance(d, dict) or d.get("archived"):
                    continue
                add(ws_res, prefix(tid, fid) + [
                    iname(key, d), d.get("unit", ""), d.get("input_value", ""),
                    d.get("progress", ""), d.get("baseline_value", ""),
                    d.get("year_baseline", ""), d.get("target_value", ""),
                    d.get("year_target", ""), d.get("level_of_result", ""),
                    d.get("data_collection", "")] + summary(tid, fid) + portfolio(tid))
        elif field not in BLOB_FIELDS and not field.startswith(ENTRY_PREFIXES):
            # scalar / narrative fields -> readable Grant Info sheet. The
            # numbered entries are skipped here; they have their own sheets
            # (and are still in All Data raw, exactly as stored).
            shown = region.get(str(value), value) if field == "region_id" else value
            add(ws_info, [tfnum, grant, tf.get(tid, {}).get("ttl", ""), year,
                          field, shown])

    # Lending Operations / Collaboration rows, grant by grant.
    def sort_key(key):
        return (str(gname(key[0])), str(fy.get(key[1], "")))

    for tid, fid in sorted(set(ops) | set(cpfs), key=sort_key):
        for n, o in enumerate(ops.get((tid, fid), []), start=1):
            add(ws_ops, prefix(tid, fid) + [
                "Lending Operation", n, o.get("informed_operation", ""),
                o.get("p_number", ""), o.get("p_code_instrument", ""),
                o.get("p_code_instrument_description", ""), o.get("approval_fy", ""),
                o.get("operation_name", ""), number(o.get("total_commitment")),
                number(o.get("informed_by_f4d")), "", "", o.get("evidence", "")])
        for n, c in enumerate(cpfs.get((tid, fid), []), start=1):
            add(ws_ops, prefix(tid, fid) + [
                "CPF", n, c.get("informed_cpf", ""), "", "", "", "", "", "", "",
                c.get("country", ""), c.get("year", ""), c.get("evidence", "")])

    for tid, fid in sorted(set(collabs) | set(collab_answer), key=sort_key):
        answer = collab_answer.get((tid, fid), "")
        entries = collabs.get((tid, fid), [])
        if not entries:
            add(ws_collab, prefix(tid, fid) + [answer, "", "", "", "", ""])
        for n, b in enumerate(entries, start=1):
            add(ws_collab, prefix(tid, fid) + [
                answer, n, b.get("type", ""), b.get("partner_detail", ""),
                b.get("describe", ""), b.get("lessons_learned", "")])

    # column widths by header name (robust to the added context columns)
    WIDE = {"Grant Name": 34, "TTL": 20, "Strategic Objective": 48, "Description": 45,
            "Next Steps": 34, "Explanation": 40, "Deliverable": 34, "Indicator": 34,
            "Pillars": 30, "Cross-Cutting Themes": 28, "F4D Association": 28,
            "Country": 18, "Region": 22, "Product Line": 14, "Data Collection": 22,
            "Value": 60, "Field": 22, "Status": 13, "Last Updated": 20, "Submitted At": 20,
            "Evidence": 50, "Operation Name": 34, "Operation Names": 34,
            "Operation Evidence": 50, "CPF Evidence": 50, "Entry": 18,
            "Instrument": 20, "Instrument (Other)": 20, "Type": 30,
            "Team(s) / Organization(s)": 30, "Lessons Learned": 40,
            "Collaboration Types": 30, "Collaboration Partners": 30,
            "Collaboration Description": 45, "Collaboration Lessons Learned": 40,
            "Collaboration Reported?": 12, "Total Commitment (US$m)": 16,
            "Informed by F4D (US$m)": 16, "Lending Operations (#)": 12,
            "Operation P Numbers": 22, "Operation Instruments": 22,
            "Operation Approval FYs": 16, "CPFs (#)": 10, "CPF Countries": 20,
            "CPF Years": 12, "Collaborations (#)": 12}
    for ws in wb.worksheets:
        for cell in ws[1]:
            ws.column_dimensions[cell.column_letter].width = WIDE.get(cell.value, 14)

    ts = datetime.datetime.now().strftime("%Y-%m-%d")
    out = os.path.abspath(f"F4D_data_export_{ts}.xlsx")
    wb.save(out)
    cur.close()
    conn.close()

    print("\nDONE. Excel file created:")
    print("   " + out)
    print(f"\nSheets: Project Progress ({ws_prog.max_row - 1} grants), "
          f"Grant Info ({ws_info.max_row - 1} rows), "
          f"Deliverables ({ws_del.max_row - 1}), "
          f"Results Indicators ({ws_res.max_row - 1}), "
          f"Lending Operations ({ws_ops.max_row - 1}), "
          f"Collaboration ({ws_collab.max_row - 1}), "
          f"All Data raw ({ws_raw.max_row - 1}).")
    if pf_headings:
        print("")
        print(f"Portfolio data merged onto Deliverables and Results Indicators: "
              f"{len(pf_matched)} of {len(pf_matched) + len(pf_unmatched)} grants matched.")
        if pf_unmatched:
            shown = sorted(pf_unmatched)
            print("   No portfolio row for: " + ", ".join(shown[:10])
                  + (f" (+{len(shown) - 10} more)" if len(shown) > 10 else ""))
            print("   Those rows still export; their portfolio columns are blank.")
    print("Open that file in Excel. (Read-only — nothing in the database changed.)")


# --------------------------------------------------------------------------- #
# Running it: show what went wrong and wait, so a double-clicked window that    #
# fails stays on screen long enough to read.                                    #
# --------------------------------------------------------------------------- #
PLAIN_ENGLISH = [
    (("login failed", "not associated with a trusted"),
     "The database did not accept that username and password. Check the "
     "password (ask Caleb) and try again -- it is typed blind, so nothing "
     "appears as you type."),
    (("timed out", "timeout", "getaddrinfo", "unreachable", "refused",
      "no route", "failed to connect", "cannot connect"),
     "Could not reach the database server. This usually means the computer "
     "is not on the World Bank network -- connect to the VDI or the VPN and "
     "try again."),
    (("permission", "denied", "not have permission"),
     "The database connected but refused to read the data. Send this file to "
     "Caleb; the account may need to be granted access."),
]


def explain(error):
    """One sentence a non-technical reader can act on, for common failures."""
    text = f"{type(error).__name__}: {error}".lower()
    for needles, message in PLAIN_ENGLISH:
        if any(n in text for n in needles):
            return message
    return str(error) or type(error).__name__


def run():
    log = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       "download_error.txt")
    try:
        main()
    except KeyboardInterrupt:
        print(os.linesep + "Stopped.")
    except Exception as error:  # noqa: BLE001 - the window must not just close
        print("")
        print("-" * 70)
        print("The download did not finish.")
        print("")
        print("   " + explain(error))
        print("")
        print("-" * 70)
        print("Technical detail (send download_error.txt to Caleb):")
        print("")
        detail = traceback.format_exc()
        print(detail)  # stdout, so it stays in order with everything above
        try:
            with open(log, "w", encoding="utf-8") as fh:
                fh.write(explain(error) + os.linesep + os.linesep + detail)
            print("Saved to: " + log)
        except OSError:
            pass
    try:
        input(os.linesep + "Press Enter to close this window ...")
    except (EOFError, KeyboardInterrupt):
        pass


if __name__ == "__main__":
    run()
